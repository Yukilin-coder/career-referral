#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LinkedIn 校友爬取脚本 - 中国人民大学 AI 校友
策略：搜索AI公司 + 中国人民大学，提取在AI公司工作的人大校友
"""

import time
import os
import sys
import re
from datetime import datetime

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def save_excel(data, filename):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RUC AI Alumni"

    headers = ["序号", "姓名", "关系", "当前职位", "当前公司", "教育背景", "LinkedIn链接", "备注"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, item in enumerate(data, 1):
        ws.append([
            i, item.get("name", ""), item.get("relation", ""),
            item.get("title", ""), item.get("company", ""),
            item.get("education", ""), item.get("link", ""), item.get("note", "")
        ])

    for col, width in zip(['A','B','C','D','E','F','G','H'], [8,18,12,35,25,30,45,20]):
        ws.column_dimensions[col].width = width

    wb.save(filename)
    print(f"Saved: {os.path.abspath(filename)}")


def is_ai_related(text):
    keywords = [
        "AI", "人工智能", "artificial intelligence", "machine learning", "深度学习",
        "算法", "推荐", "搜索", "大模型", "LLM", "AIGC", "数据科学", "Data Science",
        "ByteDance", "字节", "Baidu", "百度", "Alibaba", "阿里", "Tencent", "腾讯",
        "Moonshot", "月之暗面", "Zhipu", "智谱", "MiniMax", "Baichuan", "百川",
        "OpenAI", "Claude", "GPT", "NLP", "计算机视觉", "CV", "数据挖掘",
        "Data Mining", "算法工程师", "AI产品经理", "AI研究员", "研究员",
        "Data", "Analytics", "Quant", "量化", "模型", "Modeling"
    ]
    t = text.lower()
    return any(k.lower() in t for k in keywords)


def is_logged_in(page):
    try:
        url = page.url
        if "feed" in url or "/in/" in url:
            return True
        feed_el = page.query_selector('[data-test-id="feed-tab-icon"]')
        if feed_el:
            return True
        search = page.query_selector('input[placeholder*="Search"]') or page.query_selector('input[aria-label*="Search"]')
        if search:
            return True
        signin = page.query_selector('text=Sign in') or page.query_selector('text=Join now') or page.query_selector('text=登录')
        if signin:
            return False
    except:
        pass
    return False


def clean_person(p):
    """Python层面清理提取到的人员数据"""
    name = p.get("name", "")
    title = p.get("title", "")
    company = p.get("company", "")

    # 清理name中的关系度和认证徽章
    name = re.sub(r'\s*[·•]\s*\d+(?:nd|rd|st|th)?\+?\s*$', '', name, flags=re.I).strip()
    name = re.sub(r'\s*[·•]\s*\d+\s*度\+?\s*$', '', name).strip()
    name = re.sub(r'\s*[✓√]\s*$', '', name).strip()

    # 如果title是关系度，清空
    if re.match(r'^[·•]\s*\d+(?:nd|rd|st|th)?\+?\s*$', title, re.I):
        title = ""
    if re.match(r'^[·•]\s*\d+\s*度\+?\s*$', title):
        title = ""

    # 如果title是地点，清空
    location_patterns = [
        r'^(beijing|shanghai|shenzhen|guangzhou|hangzhou|nanjing|chengdu|wuhan|xian|tianjin)',
        r'^(北京|上海|深圳|广州|杭州|南京|成都|武汉|西安|天津|中国|美国|英国|海淀区|朝阳区|东城区|西城区|浦东新区|南山区)',
        r'^(mountain view|new york|london|san francisco|seattle|boston|area|china|united states|california)',
        r'^(haidian|chaoyang|dongcheng|xicheng|pudong|nanshan|futian|luohu|minhang|songjiang)',
    ]
    for pat in location_patterns:
        if re.match(pat, title, re.I) and len(title) < 50:
            title = ""
            break

    # 如果company是地点，清空
    for pat in location_patterns:
        if re.match(pat, company, re.I) and len(company) < 50:
            company = ""
            break

    # 如果title包含"Current:"，取后面的内容
    if title.lower().startswith('current:'):
        title = title[8:].strip()

    # 如果title为空但company有值，交换过来
    if not title and company:
        title = company
        company = ""

    p["name"] = name
    p["title"] = title
    p["company"] = company
    return p


def extract_people(page):
    """从LinkedIn people搜索结果页面提取人员信息"""
    batch = page.evaluate("""
        () => {
            const results = [];
            const seenLinks = new Set();

            const allLinks = document.querySelectorAll('a[href*="/in/"]');
            const containers = new Set();
            allLinks.forEach(link => {
                if (link.href.includes('/company/')) return;
                let c = link.closest('li');
                if (!c) c = link.closest('div[class*="search-result"], div[class*="entity-result"], div[class*="lockup"]');
                if (!c) {
                    let p = link.parentElement;
                    for (let i = 0; i < 5 && p; i++) {
                        if (p.querySelector('a[href*="/in/"]') && p.innerText.length > 40) {
                            c = p; break;
                        }
                        p = p.parentElement;
                    }
                }
                if (c) containers.add(c);
            });

            Array.from(containers).forEach(item => {
                try {
                    const linkEl = item.querySelector('a[href*="/in/"]');
                    if (!linkEl) return;
                    const link = linkEl.href.split('?')[0];
                    if (seenLinks.has(link) || link.includes('/company/')) return;
                    seenLinks.add(link);

                    const lines = item.innerText.split(String.fromCharCode(10))
                        .map(t => t.trim())
                        .filter(t => t.length > 0 && t.length < 200);

                    const isNoise = t => {
                        if (/^Connect$/i.test(t)) return true;
                        if (/^Message$/i.test(t)) return true;
                        if (t === '\u5173\u6ce8' || t === 'Follow') return true;
                        if (t === '\u9884\u7ea6') return true;
                        if (/are mutual connections/.test(t)) return true;
                        if (/is a mutual connection/.test(t)) return true;
                        if (/\u548c\u5176\u4ed6.*\u4f4d\u5171\u540c\u597d\u53cb/.test(t)) return true;
                        return false;
                    };
                    const cleanLines = lines.filter(t => !isNoise(t));

                    let name = linkEl.innerText.trim();
                    if (!name || name.length > 50) {
                        name = cleanLines[0] || '';
                    }

                    let title = '';
                    let company = '';
                    let startIdx = 0;
                    for (let i = 0; i < cleanLines.length; i++) {
                        if (cleanLines[i] === name || cleanLines[i].includes(name)) {
                            startIdx = i + 1;
                            break;
                        }
                    }

                    for (let i = startIdx; i < cleanLines.length; i++) {
                        const t = cleanLines[i];
                        if (t.toLowerCase().startsWith('current:')) continue;

                        if (!title && t.length > 2 && t.length < 120) {
                            title = t;
                        } else if (title && !company && t.length > 1 && t.length < 80 && t !== title) {
                            company = t;
                            break;
                        }
                    }

                    if (title && !company) {
                        const atMatch = title.match(/(.+?)\\s+(?:at|@|\\u5728)\\s+(.+)/i);
                        if (atMatch) {
                            title = atMatch[1].trim();
                            company = atMatch[2].trim();
                        }
                    }

                    if (name && name.length > 0 && name.length < 50) {
                        results.push({name, title, company, link});
                    }
                } catch (e) {}
            });

            return results;
        }
    """)
    return [clean_person(p) for p in batch]


def main():
    from playwright.sync_api import sync_playwright

    auth_file = "linkedin_auth.json"
    use_existing = os.path.exists(auth_file)

    with sync_playwright() as p:
        if use_existing:
            print("Loading saved login...")
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                storage_state=auth_file,
                viewport={"width": 1920, "height": 1080}
            )
            page = context.new_page()
        else:
            print("Browser window will open. Please login to LinkedIn.")
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(viewport={"width": 1920, "height": 1080})
            page = context.new_page()
            page.goto("https://www.linkedin.com/login")
            for i in range(30):
                time.sleep(2)
                if is_logged_in(page):
                    print(f"Login detected at {i*2}s")
                    break
            else:
                print("Timeout. Login not detected.")
                browser.close()
                return
            context.storage_state(path=auth_file)
            print("Login saved.")

        all_people = []
        seen = set()

        # 策略：搜索AI/科技公司 + 中国人民大学
        search_queries = [
            "ByteDance Renmin University of China",
            "Baidu 中国人民大学",
            "Tencent 中国人民大学",
            "Alibaba 中国人民大学",
            "字节跳动 中国人民大学",
            "百度 中国人民大学",
        ]

        for query in search_queries:
            if len(all_people) >= 40:
                break

            print(f"\n--- Search: '{query}' ---")
            encoded = query.replace(' ', '%20')
            url = f"https://www.linkedin.com/search/results/people/?keywords={encoded}&origin=GLOBAL_SEARCH_HEADER"
            page.goto(url, timeout=60000)
            time.sleep(5)
            page.screenshot(path=f"ruc_search_{query.replace(' ', '_')[:20]}.png")

            for i in range(10):
                batch = extract_people(page)
                new_count = 0
                for person in batch:
                    name = person.get("name", "").strip()
                    if name and len(name) < 50 and name not in seen:
                        seen.add(name)
                        all_people.append(person)
                        new_count += 1

                if new_count > 0:
                    print(f"  Scroll {i+1}: +{new_count}, total: {len(all_people)}")
                    for p in batch[:2]:
                        print(f"    -> {p.get('name')} | {p.get('title')} | {p.get('company')}")

                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(2)
                if len(all_people) >= 40:
                    break

        print(f"\nTotal extracted: {len(all_people)}")

        # 标记AI相关
        ai_people = []
        other_people = []
        for p in all_people:
            text = f"{p.get('title','')} {p.get('company','')}"
            is_ai = is_ai_related(text)
            p["relation"] = "校友"
            p["education"] = "中国人民大学"
            p["note"] = "AI相关" if is_ai else ""
            if is_ai:
                ai_people.append(p)
            else:
                other_people.append(p)

        print(f"AI related: {len(ai_people)}")
        print(f"Other: {len(other_people)}")

        # 输出：AI相关优先，补足到20人
        target = (ai_people + other_people)[:20]
        fname = f"ruc_ai_alumni_{datetime.now().strftime('%Y%m%d')}.xlsx"
        save_excel(target, fname)

        print("\nPreview (top 10):")
        for i, p in enumerate(target[:10], 1):
            print(f"{i}. {p['name']} | {p['title']} | {p['company']} | {p['note']}")

        browser.close()
        print("\nDone!")


if __name__ == "__main__":
    main()
