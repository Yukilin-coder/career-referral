import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# 模拟数据 - 基于真实爬取模式
data = [
    {"name": "张思远", "title": "Senior Talent Acquisition", "company": "ByteDance", "link": "https://www.linkedin.com/in/siyuan-zhang-demo/", "note": "HR-优先联系"},
    {"name": "李婉清", "title": "HR Business Partner", "company": "Tencent", "link": "https://www.linkedin.com/in/wanqing-li-demo/", "note": "HR-优先联系"},
    {"name": "王浩宇", "title": "Recruiting Manager", "company": "Baidu", "link": "https://www.linkedin.com/in/haoyu-wang-demo/", "note": "HR-优先联系"},
    {"name": "刘子涵", "title": "Data Scientist", "company": "ByteDance", "link": "https://www.linkedin.com/in/zihan-liu-demo/", "note": "岗位匹配-优先"},
    {"name": "陈嘉明", "title": "Algorithm Engineer", "company": "Baidu", "link": "https://www.linkedin.com/in/jiaming-chen-demo/", "note": "岗位匹配-优先"},
    {"name": "赵雨桐", "title": "Senior ML Engineer", "company": "Alibaba", "link": "https://www.linkedin.com/in/yutong-zhao-demo/", "note": "岗位匹配-优先"},
    {"name": "周星辰", "title": "AI Product Manager", "company": "Tencent", "link": "https://www.linkedin.com/in/xingchen-zhou-demo/", "note": "岗位匹配-优先"},
    {"name": "吴宇航", "title": "NLP Researcher", "company": "ByteDance", "link": "https://www.linkedin.com/in/yuhang-wu-demo/", "note": "岗位匹配-优先"},
    {"name": "孙雅琳", "title": "Data Analyst", "company": "Alibaba", "link": "https://www.linkedin.com/in/yalin-sun-demo/", "note": "岗位匹配-优先"},
    {"name": "马俊杰", "title": "推荐算法工程师", "company": "Baidu", "link": "https://www.linkedin.com/in/junjie-ma-demo/", "note": "岗位匹配-优先"},
    {"name": "林晓彤", "title": "Marketing Manager", "company": "ByteDance", "link": "https://www.linkedin.com/in/xiaotong-lin-demo/", "note": "校友"},
    {"name": "黄思琪", "title": "Corporate Strategy", "company": "Tencent", "link": "https://www.linkedin.com/in/siqi-huang-demo/", "note": "校友"},
    {"name": "郑浩然", "title": "Sales Director", "company": "Baidu", "link": "https://www.linkedin.com/in/haoran-zheng-demo/", "note": "校友"},
    {"name": "何梦瑶", "title": "Finance Analyst", "company": "Alibaba", "link": "https://www.linkedin.com/in/mengyao-he-demo/", "note": "校友"},
    {"name": "谢梓轩", "title": "Operations Manager", "company": "ByteDance", "link": "https://www.linkedin.com/in/zixuan-xie-demo/", "note": "校友"},
    {"name": "冯诗涵", "title": "Legal Counsel", "company": "Tencent", "link": "https://www.linkedin.com/in/shihan-feng-demo/", "note": "校友"},
    {"name": "许明哲", "title": "UX Designer", "company": "Baidu", "link": "https://www.linkedin.com/in/mingzhe-xu-demo/", "note": "校友"},
    {"name": "曹欣怡", "title": "Content Strategist", "company": "Alibaba", "link": "https://www.linkedin.com/in/xinyi-cao-demo/", "note": "校友"},
    {"name": "彭凯文", "title": "BD Manager", "company": "ByteDance", "link": "https://www.linkedin.com/in/kaiwen-peng-demo/", "note": "校友"},
    {"name": "邓若曦", "title": "PMO", "company": "Tencent", "link": "https://www.linkedin.com/in/ruoxi-deng-demo/", "note": "校友"},
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Referral Contacts"

headers = ["序号", "姓名", "关系", "当前职位", "当前公司", "教育背景", "LinkedIn链接", "备注"]
ws.append(headers)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

for i, item in enumerate(data, 1):
    ws.append([
        i,
        item["name"],
        "校友",
        item["title"],
        item["company"],
        "中国人民大学",
        item["link"],
        item["note"],
    ])

# 设置列宽
for col, width in zip(['A','B','C','D','E','F','G','H'], [8, 15, 10, 30, 18, 15, 50, 18]):
    ws.column_dimensions[col].width = width

# 根据备注设置行颜色
for row_idx, item in enumerate(data, 2):
    note = item["note"]
    if "HR" in note:
        fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    elif "岗位匹配" in note:
        fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    else:
        fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    for col_idx in range(1, 9):
        ws.cell(row=row_idx, column=col_idx).fill = fill

output_path = os.path.join(os.path.dirname(__file__), "referral-contacts-demo.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
